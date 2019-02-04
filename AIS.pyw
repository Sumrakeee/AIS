import os
import re
import sys
from PyQt5 import QtWidgets, QtGui
from PyQt5.QtWidgets import (QMainWindow, 
	QTextEdit, QAction, QFileDialog, QApplication, QMessageBox)
from PyQt5.QtGui import QIcon
from frontend import Ui_AIS
from openpyxl import Workbook
from openpyxl import load_workbook

class Functions:
	# Тестовая функция для отображения большого количества строк в целях проверки скроллбокса
	def getList(self, path='C:\\Windows\\System32'):
		ld = os.listdir(path)
		stringlist = ''

		for i in ld:
			stringlist += ''.join(i) + '\n'

		return stringlist

class MainWindow(QtWidgets.QMainWindow):
	def __init__(self):
		super(MainWindow, self).__init__()
		self.ui = Ui_AIS()
		self.ui.setupUi(self)
		# Фиксирует размер окна и запрещает максимизацию
		self.setFixedSize(self.size()) 

		# # # Связи

		openFile = QAction('Открыть файл', self)
		openFile.setShortcut('Ctrl+O')
		openFile.triggered.connect(self.OpenFileDialog)

		menubar = self.menuBar()
		fileMenu = menubar.addMenu('&Файл')
		fileMenu.addAction(openFile)


		self.ui.ButtonSearch.clicked.connect(self.Search)
		self.ui.ButtonExcelExport.clicked.connect(self.ExportFileDialog)

		self.ui.LabelSearchResult.setText(Functions.getList(self))

	# # # Функции

	def Search(self): # !!! Внедрить поисковый алгоритм using re или BeautifulSoup
		searchText = self.ui.LineSub.text() + ':' + self.ui.LineRay.text() + ':' + self.ui.LineKva.text() + ':' + self.ui.LineObn.text()

		self.ui.LabelSearchResult.setText(searchText)

		self.ui.LabelProps.setText('<a href="http://google.com">Hyperlink example</a>')
		self.ui.LabelProps.setOpenExternalLinks(True)

	def ExcelExport(self, filename):
		props = self.ui.LabelProps.text()
		
		# В словарь заносятся данные по характеристикам с ключом
		data = {'props': props} 

		# Проверка на наличие файла с необходимым именем
		try:
			file = open(filename)
		except IOError as e:
			wb = Workbook()
			ws = wb.active

			# Заголовки колонок (не переменные, задаются статически строками)
			wr = ( (
				'Характеристика 1',
				'Характеристика 2') )

			ws.append(wr)
			wb.save(filename)

		book = load_workbook(filename)
		sheet = book.active

		# Запись данных из словаря в колонки по ключу
		row = ( (
			data['props'],
			data['props']) )

		sheet.append(row)
		book.save(filename)

	def OpenFileDialog(self):
		# Открывает окно выбора файла и записывает путь в переменную
		fname = QFileDialog.getOpenFileName(self, 'Открыть')[0]

		if fname != '':
			os.startfile(fname)
		else:
			pass


	def ExportFileDialog(self):
		# Открывает окно выбора файла и записывает путь в переменную
		fname = QFileDialog.getOpenFileName(self, 'Экспорт в Excel')[0]

		if fname != '':
			try:
				# Вызов функции экспорта в Excel
				self.ExcelExport(fname)
				# Открывает окно с сообщением об удачном экспорте
				reply = QMessageBox.information(self, 'Экспорт в Excel', 'Экспортровано в %s' % fname, QMessageBox.Ok)
			except Exception as e:
				# Открывает окно с сообщением об ошибке и ее описанием
				reply = QMessageBox.critical(self, 'Ошибка', 'Ошибка экспорта в Excel\n\n%s' % str(e).replace('openpyxl', 'Exporter'), QMessageBox.Ok)
		else:
			pass


if __name__ == '__main__':
	app = QtWidgets.QApplication([])
	application = MainWindow()
	application.show()

	sys.exit(app.exec())